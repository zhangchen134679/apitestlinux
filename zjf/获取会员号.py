import random

import openpyxl
import requests
from requests.auth import HTTPBasicAuth
import time
import re
import json
from openpyxl.worksheet.worksheet import Worksheet

# 获取token
urlToken = 'https://dl-api.lorealchina.com/api/interface/oauth/token'
login = {
	'grant_type': 'password',
	'username': 'crm3ce',
	'password': 'bvjd1jzldijnz4e0'
	}
r = requests.post(url=urlToken, json=login, auth=HTTPBasicAuth("crm_3ce", "kiaa4xym"))

print(r.json())

#加工token
bearToken = 'Bearer' + ' ' + r.json()["access_token"]

#注册用户
url = 'https://dl-api.lorealchina.com/api/interface/third/vb/member/register'

value2 = '139'+str(random.randint(10000000, 90000000))
mobile = json.loads(value2)

newMember = {
    "brand_code": "3ce",
    "program_code": "3ce",
    "name":"abcdefghijk!@#lmnopq!@#rstuvwxyz",
    "mobile":mobile,
    "chanel_code": "9AD",
    "consentStatus": "1",
    "consentTime": "2020-10-27 13:00:00",
    "store_code": "13660381G1"
}

#生成ruid
CurrentTime = time.time()
ruid = '{' + '"uuid":"{a}"'.format(a=CurrentTime) + '}'

header = {
    "Content-Type": "application/json",
    "Authorization": bearToken,
    "ruid": ruid
}
for r in range(1, 13):
    r = requests.post(url=url, json=newMember, headers=header)
    #da = r.json()['data']['cardno']
    print(r.json())
#card = da.json()['cardno']


#print(card)

#work_book = openpyxl.load_workbook("cases.xlsx")
#sheet = work_book["Sheet1"]
#for i in range(1,100001):
#    sheet.cell(i, 1).value = "test1111"
#work_book.save("cases.xlsx")
#work_book.close()

#def write(sheet_name, data1, data2, data3, path):

#循环写入excel
work_book = openpyxl.load_workbook('ba上传模板.xlsx')
sheet: Worksheet = work_book['BA']
for n in range(2, 13):
        sheet.cell(n, 1).value = da
        sheet.cell(n, 2).value = 'Test001'
        sheet.cell(n, 3).value = 'Test002'
work_book.save('ba上传模板.xlsx')
work_book.close()
#if __name__ == '__main__':
#    write("BA", '', 'Test001', 'Test002', 'ba上传模板.xlsx')
print('end')
