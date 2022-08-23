import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
def write(sheet_name, data1, data2, data3, path):
    work_book = openpyxl.load_workbook(path)
    sheet: Worksheet = work_book[sheet_name]
    for n in range(2, 100002):
        sheet.cell(n, 1).value = data1
        sheet.cell(n, 2).value = data2
        sheet.cell(n, 3).value = data3
    work_book.save(path)
    work_book.close()
if __name__ == '__main__':
    write("BA", '', 'Test001', 'Test002', 'ba上传模板.xlsx')
print('end')
